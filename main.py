import os, requests
from dotenv import load_dotenv
from datetime import datetime
import csv
from send_email import send_weather_report
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from fpdf import FPDF

def ensure_csv_exists(filepath, fieldnames):
    if not os.path.isfile(filepath):
        with open(filepath, 'w', encoding='utf-8') as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()


def append_csv_row(filepath, data_dict):
    with open(filepath, 'a', encoding='utf-8') as file:
        fieldnames = list(data_dict.keys())
        writer = csv.DictWriter(file, fieldnames)
        writer.writerow(data_dict)

def fetch_weather(city):
# Monta a URL
    KEY = os.getenv('OPENWEATHER_API_KEY')
    URL = f'https://api.openweathermap.org/data/2.5/weather?q={city}&appid={KEY}&units=metric&lang=pt_br'

    # Fazer o request e armazena o resultado em result_dict
    try:
        response = requests.get(URL)
        data = response.json()
        if not response.ok:
            return {
                'ok': False,
                'error': data.get('message', 'Unknown error'),
                'status_code': response.status_code
            }
        
        result_dict = {
                'ok': True,
                'data':{
                    'datetime': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                    'city': data['name'],
                    'temperature': data['main']['temp'],
                    'humidity': data['main']['humidity'],
                    'description': data['weather'][0]['description']
                }
            }
    except Exception as e:
        result_dict = {
            'ok': False,
            'error': 'Error',
            'status_code': 500
        }
    
    return result_dict


# FASE A — Dados para Relatório
def read_csv_rows(filepath):
    with open(filepath, 'r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        rows = []
        for line in reader:
            rows.append(line)

    for line in rows:
        line['datetime'] = datetime.strptime(line['datetime'], '%d/%m/%Y %H:%M:%S')

    rows.sort(key=lambda line: line['datetime'])
    return rows

# FASE B — Criar Excel do Relatório
def generate_xlsx_report(rows, fieldnames):
    folder_name = datetime.now().strftime("%Y-%m-%d")
    folder_path = os.path.join('reports', folder_name)
    os.makedirs(folder_path, exist_ok=True)

    # 1. Criar workbook e worksheet
    report_wb = Workbook()
    report_ws = report_wb.active
    # 2. Escrever cabeçalhos
    for pos, header in enumerate(fieldnames, start=1):
        report_ws.cell(row=1, column=pos).value = header
    # 3. Preencher dados ordenados
    linha_atual = 2
    for line in rows:
        for pos, header in enumerate(fieldnames, start=1):
            report_ws.cell(row=linha_atual, column=pos).value = line[header]
        linha_atual += 1

    # 4. Estilizar cabeçalhos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    medium_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

    for pos in range(1, len(fieldnames) + 1):
        cell = report_ws.cell(row=1, column=pos)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = medium_border

    # 5. Estilizar dados (alinhamento, bordas, zebra)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    fill_odd = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    fill_even = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')

    align_left = Alignment(horizontal='left', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')

    for row in range(2, report_ws.max_row + 1):
        for col in range(1, len(fieldnames) + 1):
            cell = report_ws.cell(row=row, column=col)
            cell.border = thin_border

            # Escolhe alinhamento com base no nome da coluna
            header = fieldnames[col - 1]
            if header in ['city', 'description']:
                cell.alignment = align_left
            else:
                cell.alignment = align_center

            # Linhas zebradas
            if row % 2 == 0:
                cell.fill = fill_even
            else:
                cell.fill = fill_odd


    # 6. Ajustar largura das colunas
    for col in range(1, len(fieldnames) + 1):
        max_length = 0
        for row in range(1, report_ws.max_row + 1):
            cell = report_ws.cell(row=row, column=col)
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        col_letter = get_column_letter(col)
        report_ws.column_dimensions[col_letter].width = max_length + 2 

    # 7. Congelar linha do cabeçalho
    report_ws.freeze_panes = 'A2'
    # 8. Ativar filtros nas colunas
    report_ws.auto_filter.ref = report_ws.dimensions
    # 9. Salvar como "weather_report.xlsx"
    report_wb.save(os.path.join(folder_path, 'weather_report.xlsx'))
    return os.path.join(folder_path, 'weather_report.xlsx')


# FASE C — Criar PDF do Relatório

def create_pdf_report(rows):
    folder_name = datetime.now().strftime("%Y-%m-%d")
    folder_path = os.path.join('reports', folder_name)
    os.makedirs(folder_path, exist_ok=True)

    # 1 - Criar o objeto PDF
    pdf = FPDF()
    # 2 - Criar uma página
    pdf.add_page()
    # 3 - Definir fonte
    pdf.set_font(family='Arial', size=16)
    # 4 - Escrever o título do relatório
    pdf.cell(0, 14, 'Weather Report - Consultation History', ln=1, align='C')
    # 5 - Escrever data/hora de geração
    pdf.set_font(family='Arial', size=10)
    pdf.cell(0, 10, f'Generated in: {datetime.now().strftime("%d/%m/%Y %H:%M")}', ln=1)
    pdf.ln(5)
    # 6 - Escrever o resumo:
    # 6.1 - Total de registros no histórico
    pdf.cell(0, 10, f'Total records: {len(rows)}', ln=1)
    # 6.2 - Data da primeira medição
    first = rows[0]
    pdf.cell(0, 10, f'First consultation: {first["datetime"].strftime("%d/%m/%Y - %H:%M")}', ln=1)
    # 6.3 - Data da última medição
    last = rows[-1]
    pdf.cell(0, 10, f'Last consultation: {last["datetime"].strftime("%d/%m/%Y - %H:%M")}', ln=1)
    pdf.ln(4)
    # 6.4 - Últimas 3 medições (datetime, city, temp, description)
    last_consultations = []
    if len(rows) >= 1:
        last_consultations.append(rows[-1])
    if len(rows) >= 2:
        last_consultations.append(rows[-2])
    if len(rows) >= 3:
        last_consultations.append(rows[-3])

    pdf.set_font(family='Arial', size=14)
    pdf.cell(0, 10, 'Last three consultations', ln=1, align='C')
    pdf.set_font(family='Arial', size=10)

    for item in last_consultations:
        text = f'{item["datetime"].strftime("%d/%m/%Y - %H:%M")} | {item["city"]}'
        text2 = f'{item["temperature"]}°C | {item["humidity"]}% | {item["description"]}'
        pdf.cell(0, 10, text, ln=1 )
        pdf.cell(0, 10, text2, ln=1)
        pdf.ln(4)
    # 7 - Salvar arquivo "weather_report.pdf"
    filepath = os.path.join(folder_path, 'weather_report.pdf')
    pdf.output(filepath)
    return filepath


def run_automation():
    load_dotenv()
    city = os.getenv("DEFAULT_CITY")
    filepath = 'weather_log.csv'
    fieldnames = ['datetime', 'city', 'temperature', 'humidity', 'description']
    ensure_csv_exists(filepath=filepath, fieldnames=fieldnames)
    result = fetch_weather(city=city)
    if not result['ok']:
        msg = result['error']
        status = result['status_code']
        print(f'Fetch failed: <{msg}> ({status})')
        return False

    data = result['data']
    append_csv_row(filepath, data)
    rows = read_csv_rows(filepath)
    xlsx_path = generate_xlsx_report(rows, fieldnames)
    pdf_path = create_pdf_report(rows)
    send_weather_report(result=data, pdf_path=pdf_path, xlsx_path=xlsx_path)
    return True

if __name__ == "__main__":
    run_automation()
